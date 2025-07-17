import re
import io
from typing import Dict, List, Any
from openpyxl import load_workbook

class LatexGenerator:
    def __init__(self):
        pass

    def sanitize_filename(self, name: str) -> str:
        """Sanitize filename for safe file creation."""
        sanitized = re.sub(r'[<>:"/\\|?*]', '_', str(name))
        sanitized = re.sub(r'\s+', '_', sanitized.strip())
        return sanitized

    def escape_latex(self, text: Any) -> str:
        """Escape special LaTeX characters."""
        if not text:
            return ""

        text_str = str(text)

        def _escape_chars(s: str) -> str:
            return (
                s.replace("\\", "\\textbackslash{}")
                .replace("&", "\\&")
                .replace("%", "\\%")
                .replace("$", "\\$")
                .replace("#", "\\#")
                .replace("_", "\\_")
                .replace("{", "\\{")
                .replace("}", "\\}")
                .replace("~", "\\textasciitilde{}")
                .replace("^", "\\textasciicircum{}")
            )

        if "•" in text_str:
            text_str = text_str.replace("\n", " ")
            items = [item.strip() for item in text_str.split("•") if item.strip()]
            escaped_items = ["• " + _escape_chars(item) for item in items]
            return " \\newline ".join(escaped_items)
        else:
            return _escape_chars(text_str)

    def col_letter_to_index(self, letter: str) -> int:
        """Convert Excel column letter to zero-based index."""
        return ord(letter.upper()) - ord("A")

    def process_worksheet(self, ws, sheet_name: str, config: Dict[str, Any]) -> str:
        """
        Process a single worksheet and generate LaTeX table.
        
        Supports two column width modes:
        
        1. **Manual Mode** (default):
           - Use config['column_widths'] to specify individual column widths
           - Example: {"A": "6.7", "B": "4.0", "C": "2.0"}
           
        2. **Auto-Distribution Mode**:
           - Use config['total_table_width'] to specify total table width
           - System automatically distributes width among included columns
           - Maintains 3:1 ratio (A/B columns get triple width of others)
           - Example: {"total_table_width": 15.6}
           
        Args:
            ws: Excel worksheet object
            sheet_name: Name of the worksheet
            config: Configuration dictionary with options:
                - excluded_columns: List of column letters to exclude (default: ["B", "C", "D", "E"])
                - column_widths: Manual column widths (used if total_table_width not provided)
                - total_table_width: Total width for auto-distribution (overrides manual widths)
                - font_size: LaTeX font size command ('footnotesize' or 'scriptsize', default: 'footnotesize')
        
        Returns:
            LaTeX table string
        """
        excluded_columns = set(config.get('excluded_columns', ["B", "C", "D", "E"]))
        
        total_table_width = config.get('total_table_width')
        use_auto_distribution = total_table_width is not None
        
        all_letters = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
        final_letters = [l for l in all_letters if l not in excluded_columns]
        final_indices = [self.col_letter_to_index(l) for l in final_letters]
        
        if use_auto_distribution:
            total_width = float(total_table_width) if total_table_width else 15.6
            
            n_AB = len([l for l in final_letters if l in {"A", "B"}])
            n_others = len([l for l in final_letters if l not in {"A", "B"}])
            
            # Auto-distribution formula with 3:1 ratio: w_others = total / (3 * n_AB + n_others)
            if (3 * n_AB + n_others) > 0:
                w_others = total_width / (3 * n_AB + n_others)
                w_AB = 3 * w_others
            else:
                w_others = w_AB = total_width / len(final_letters) if final_letters else 1.0
            
            COLUMN_WIDTHS = {}
            for letter in final_letters:
                if letter in {"A", "B"}:
                    COLUMN_WIDTHS[letter] = f"m{{{w_AB:.2f}cm}}"
                else:
                    COLUMN_WIDTHS[letter] = f">{{\\centering\\arraybackslash}}m{{{w_others:.2f}cm}}"
        else:
            widths = config.get('column_widths', {})
            COLUMN_WIDTHS = {
                "A": f"m{{{widths.get('A', '6.7')}cm}}",
                "B": f"m{{{widths.get('B', '4.0')}cm}}",
                "C": f">{{\\centering\\arraybackslash}}m{{{widths.get('C', '2.0')}cm}}",
            }

        excluded_indices = set(self.col_letter_to_index(c) for c in excluded_columns)
        header_title = self.escape_latex(ws["A1"].value) if ws["A1"].value else "Table"

        headers = []
        for letter in final_letters:
            cell_ref = f"{letter}2"
            header_value = ws[cell_ref].value if ws[cell_ref].value else f"Column {letter}"
            headers.append(self.escape_latex(header_value))

        total_width = 0.0
        for l in final_letters:
            if use_auto_distribution:
                width_str = COLUMN_WIDTHS[l]
                import re
                width_match = re.search(r'm\{([\d.]+)cm\}', width_str)
                if width_match:
                    width_value = width_match.group(1)
                else:
                    width_value = '2.0'
            else:
                if l in {"A", "B"}:
                    width_value = widths.get(l, '4.0')
                else:
                    width_value = widths.get('C', '2.0')
            total_width += float(width_value)
        
        total_width_str = f"{total_width:.2f}cm"

        tabular_parts = []
        for l in final_letters:
            if use_auto_distribution:
                tabular_parts.append(COLUMN_WIDTHS[l])
            else:
                if l in {"A", "B"}:
                    tabular_parts.append(COLUMN_WIDTHS[l])
                else:
                    tabular_parts.append(COLUMN_WIDTHS["C"])
        tabular_spec = "|" + "|".join(tabular_parts) + "|"

        color_definition = """% Add these packages to your LaTeX document preamble:
% \\usepackage{array}
% \\usepackage{xcolor}
% \\usepackage{colortbl}
% \\usepackage{longtable}
\\definecolor{headercolor}{HTML}{00ACD2}
"""

        # Get font size from config (default: footnotesize)
        font_size = config.get('font_size', 'footnotesize')
        
        header_latex = f"""
\\{font_size}
\\begin{{longtable}}{{{tabular_spec}}}
\\hline
\\rowcolor{{headercolor}}\\multicolumn{{{len(final_letters)}}}{{|c|}}{{\\parbox[c][4ex][c]{{{total_width_str}}}{{\\centering\\textcolor{{white}}{{\\textbf{{\\normalsize{{{header_title}}}}}}}}}}} \\\\
\\hline
\\rowcolor{{headercolor}}{" & ".join([f"\\textcolor{{white}}{{\\textbf{{{header}}}}}" for header in headers])} \\\\
\\hline
\\endfirsthead

\\hline
\\rowcolor{{headercolor}}\\multicolumn{{{len(final_letters)}}}{{|c|}}{{\\parbox[c][4ex][c]{{{total_width_str}}}{{\\centering\\textcolor{{white}}{{\\textbf{{\\normalsize{{{header_title}}} (continued)}}}}}}}} \\\\
\\hline
\\rowcolor{{headercolor}}{" & ".join([f"\\textcolor{{white}}{{\\textbf{{{header}}}}}" for header in headers])} \\\\
\\hline
\\endhead

\\hline
\\multicolumn{{{len(final_letters)}}}{{|r|}}{{\\textit{{Continued on next page...}}}} \\\\
\\hline
\\endfoot

\\hline
\\endlastfoot
"""

        body_rows = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            original_filtered = [row[i] if i < len(row) else None for i in final_indices]
            row_has_content = any(
                cell is not None and str(cell).strip() 
                for cell in original_filtered
            )
            
            if not row_has_content:
                continue
                
            escaped = [self.escape_latex(cell) for cell in row]
            filtered = [escaped[i] for i in final_indices]
            
            has_content_only_in_A = False
            if "A" not in excluded_columns:
                first_col_has_content = original_filtered[0] is not None and str(original_filtered[0]).strip()
                other_cols_empty = all(
                    not (cell is not None and str(cell).strip()) 
                    for cell in original_filtered[1:]
                )
                has_content_only_in_A = first_col_has_content and other_cols_empty
            
            if has_content_only_in_A:
                body_rows.append("\\rowcolor{headercolor}" + " & ".join(filtered) + " \\")
            else:
                body_rows.append(" & ".join(filtered) + " \\")

        body_latex = "\n\\hline\n".join(body_rows)
        latex_table = header_latex + body_latex + "\n\\end{longtable}"
        
        return color_definition + "\n" + latex_table

    def process_excel(self, excel_content: bytes, config: Dict[str, Any]) -> Dict[str, str]:
        """Process entire Excel file and return LaTeX files."""
        latex_files = {}
        
        try:
            wb = load_workbook(filename=io.BytesIO(excel_content))
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                latex_content = self.process_worksheet(ws, sheet_name, config)
                filename = f"{self.sanitize_filename(sheet_name)}.tex"
                latex_files[filename] = latex_content
                
            return latex_files
            
        except Exception as e:
            raise Exception(f"Error processing Excel file: {str(e)}")